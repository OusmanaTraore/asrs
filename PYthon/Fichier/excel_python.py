
import openpyxl

wb= openpyxl.load_workbook('octobre.xlsx')
print(wb.sheetnames)

#Recuperer la feuille
# sheet = wb["Feuil1"]
sheet = wb.active

cell = sheet["B7"]
print(cell.value)

print(sheet.max_row)
print(sheet.max_column)
# cell = sheet.cell(4,3) #y(colone) ,x(row)

for row in range(2,7):
    cell = sheet.cell(row,2)
    print(cell.value)


